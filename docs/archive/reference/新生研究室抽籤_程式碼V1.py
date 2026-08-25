import pandas as pd
import random
import openpyxl
import re
import datetime
import json
import IPython

def parse_seat_label(val):
    if val is None:
        return None
    if isinstance(val, (datetime.datetime, datetime.date)):
        return f"{val.month}-{val.day}"
    s = str(val).strip()
    m = re.match(r'^(\d+)\s*[-/月]\s*(\d+)\s*日?$', s)
    if m:
        return f"{int(m.group(1))}-{int(m.group(2))}"
    return None

def run_real_floorplan_lottery(input_file="研究室選位.xlsx", output_excel="研究室選位_蛇形走位完成版.xlsx", output_html="研究室抽籤大會_實體格局版.html"):
    df = pd.read_excel(input_file, sheet_name=0)
    groups = []
    for idx, row in df.iterrows():
        group_id = row['組別']
        if pd.isna(group_id):
            continue
        members = []
        for col_val in row.iloc[3:]:
            if pd.notna(col_val) and str(col_val).strip() != "":
                members.append(str(col_val).strip())
        if len(members) > 0:
            groups.append({
                'row_index': idx,
                'group_id': int(group_id),
                'rep': members[0],
                'members': members,
                'size': len(members)
            })

    caps = {"209": 20, "310": 27, "313": 23, "919": 9}
    lab_names = ["209", "310", "313", "919"]

    shuffled_groups = [dict(g) for g in groups]
    random.shuffle(shuffled_groups)

    def solve(idx, bins, assignments):
        if idx == len(shuffled_groups):
            return True
        g = shuffled_groups[idx]
        choices = list(range(4))
        random.shuffle(choices)
        for b_i in choices:
            lab = lab_names[b_i]
            if sum(bins[b_i]) + g['size'] <= caps[lab]:
                bins[b_i].append(g['size'])
                assignments.append((g, lab))
                if solve(idx + 1, bins, assignments):
                    return True
                assignments.pop()
                bins[b_i].pop()
        return False

    bins = [[] for _ in range(4)]
    assignments = []
    if not solve(0, bins, assignments):
        print("❌ 分配失敗")
        return

    wb = openpyxl.load_workbook(input_file)

    # 4 間房間 100% 擬真格局 S 型蛇形走位序列
    snake_layouts = {
        "209": [
            "1-1","1-2","1-3","1-4","1-5",
            "2-5","2-4","2-3","2-2","2-1",
            "3-1","3-2","3-3","3-4","3-5",
            "4-5","4-4","4-3","4-2","4-1"
        ],
        "310": [
            "1-1","1-2","1-3","1-4","1-5",
            "2-5","2-4","2-3","2-2","2-1",
            "3-1","3-2","3-3","3-4","3-5",
            "4-5","4-4","4-3","4-2","4-1",
            "5-1","5-2","5-3","5-4","5-5",
            "6-2","6-1"
        ],
        "313": [
            "1-1","1-2","1-3","1-4","1-5","1-6",
            "2-6","2-5","2-4","2-3","2-2","2-1",
            "3-1","3-2","3-3","3-4","3-5","3-6",
            "4-5","4-4","4-3","4-2","4-1"
        ],
        "919": [
            "1-1","1-2","1-3","1-4",
            "2-5","2-4","2-3","2-2","2-1"
        ]
    }

    lab_counters = {lab: 0 for lab in lab_names}
    final_allocation = []
    seat_mapping_code = {}

    for g, lab in assignments:
        start_idx = lab_counters[lab]
        g_size = g['size']
        assigned_labels = snake_layouts[lab][start_idx : start_idx + g_size]
        lab_counters[lab] += g_size

        final_allocation.append({
            'group_id': g['group_id'],
            'rep': g['rep'],
            'members': g['members'],
            'size': g['size'],
            'lab': lab,
            'seat_labels': assigned_labels,
            'row_index': g['row_index']
        })

        for student_name, label in zip(g['members'], assigned_labels):
            seat_mapping_code[f"{lab}-{label}"] = student_name

    # 更新工作表 1
    ws1 = wb.worksheets[0]
    for item in final_allocation:
        excel_row = item['row_index'] + 2
        ws1.cell(row=excel_row, column=2, value=str(item['lab']))
        ws1.cell(row=excel_row, column=3, value=", ".join(item['seat_labels']))

    # 更新各研究室工作表 (右側儲存格)
    for lab_name in lab_names:
        if lab_name in wb.sheetnames:
            ws_lab = wb[lab_name]
            for row in ws_lab.iter_rows():
                for cell in row:
                    p = parse_seat_label(cell.value)
                    if p and f"{lab_name}-{p}" in seat_mapping_code:
                        ws_lab.cell(row=cell.row, column=cell.column + 1, value=seat_mapping_code[f"{lab_name}-{p}"])

    wb.save(output_excel)

    # 輸出資料供網頁使用
    draw_data = []
    for order_idx, item in enumerate(final_allocation, start=1):
        draw_data.append({
            'order': order_idx,
            'group_id': item['group_id'],
            'size': item['size'],
            'rep': item['rep'],
            'members': item['members'],
            'lab': item['lab'],
            'seat_labels': item['seat_labels']
        })

    html_code = f"""<!DOCTYPE html>
<html lang="zh-TW">
<head>
<meta charset="UTF-8">
<title>新生茶會 - 研究室實體格局開獎系統</title>
<script src="https://cdn.jsdelivr.net/npm/canvas-confetti@1.6.0/dist/confetti.browser.min.js"></script>
<style>
  :root {{
    --bg-color: #0b1120;
    --card-bg: #1e293b;
    --primary: #3b82f6;
    --accent: #f59e0b;
    --text-main: #f8fafc;
    --text-sub: #94a3b8;
    --table-border: #475569;
  }}
  * {{ box-sizing: border-box; margin: 0; padding: 0; font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, "Microsoft JhengHei", sans-serif; }}
  body {{ background-color: var(--bg-color); color: var(--text-main); padding: 20px; }}
  .container {{ max-width: 1240px; margin: 0 auto; }}
  
  header {{ text-align: center; margin-bottom: 20px; }}
  header h1 {{ font-size: 2.2rem; color: #60a5fa; margin-bottom: 6px; text-shadow: 0 0 20px rgba(96,165,250,0.4); }}
  header p {{ color: var(--text-sub); font-size: 1rem; }}

  /* Room progress banner */
  .room-banner {{ display: grid; grid-template-columns: repeat(4, 1fr); gap: 15px; margin-bottom: 20px; }}
  .room-card {{ background: var(--card-bg); border-radius: 12px; padding: 12px 16px; border: 2px solid #334155; }}
  .room-title {{ display: flex; justify-content: space-between; font-weight: bold; margin-bottom: 6px; font-size: 1.05rem; }}
  .room-progress-bg {{ background: #334155; height: 10px; border-radius: 5px; overflow: hidden; }}
  .room-progress-bar {{ height: 100%; width: 0%; background: linear-gradient(90deg, #3b82f6, #10b981); transition: width 0.4s ease; }}
  .room-stat {{ font-size: 0.85rem; color: var(--text-sub); margin-top: 5px; text-align: right; }}

  /* Stage Area */
  .stage-box {{ background: var(--card-bg); border-radius: 18px; padding: 25px 30px; text-align: center; margin-bottom: 25px; border: 1px solid #334155; box-shadow: 0 10px 30px rgba(0,0,0,0.6); }}
  .stage-title {{ font-size: 1.15rem; color: var(--accent); letter-spacing: 1.5px; font-weight: bold; margin-bottom: 12px; }}
  .group-info-header {{ font-size: 1.6rem; font-weight: 800; color: #93c5fd; margin-bottom: 15px; }}

  .members-grid-box {{ display: flex; flex-wrap: wrap; justify-content: center; gap: 12px; max-width: 1000px; margin: 0 auto 20px auto; min-height: 55px; align-items: center; }}
  .member-large-tag {{ 
    background: #0f172a; 
    border: 2px solid #475569; 
    color: #f8fafc; 
    font-size: 1.7rem; 
    font-weight: 800; 
    padding: 8px 22px; 
    border-radius: 12px; 
    display: inline-flex; 
    align-items: center; 
    box-shadow: 0 4px 10px rgba(0,0,0,0.3);
    animation: popIn 0.3s cubic-bezier(0.175, 0.885, 0.32, 1.275);
  }}
  @keyframes popIn {{
    0% {{ transform: scale(0.7); opacity: 0; }}
    100% {{ transform: scale(1); opacity: 1; }}
  }}

  .result-banner {{ display: inline-block; padding: 14px 45px; font-size: 2.6rem; font-weight: 900; border-radius: 14px; background: #0f172a; border: 3px dashed #475569; color: #64748b; margin-bottom: 18px; min-width: 320px; transition: all 0.3s; }}
  .result-banner.revealed {{ background: linear-gradient(135deg, #1e3a8a, #065f46); color: #34d399; border: 3px solid #10b981; transform: scale(1.06); box-shadow: 0 0 35px rgba(16,185,129,0.5); }}

  .btn-group {{ display: flex; justify-content: center; gap: 15px; }}
  .btn {{ padding: 12px 28px; font-size: 1.15rem; font-weight: bold; border: none; border-radius: 10px; cursor: pointer; transition: all 0.2s; }}
  .btn-primary {{ background: #3b82f6; color: white; box-shadow: 0 4px 14px rgba(59,130,246,0.4); }}
  .btn-primary:hover {{ background: #2563eb; transform: translateY(-2px); }}
  .btn-accent {{ background: #f59e0b; color: #0f172a; box-shadow: 0 4px 14px rgba(245,158,11,0.4); }}
  .btn-accent:hover {{ background: #d97706; transform: translateY(-2px); }}
  .btn:disabled {{ opacity: 0.4; cursor: not-allowed; }}

  /* Floor Plan Layouts */
  .floor-section {{ margin-top: 25px; }}
  .floor-tabs {{ display: flex; gap: 10px; margin-bottom: 15px; justify-content: center; }}
  .tab-btn {{ padding: 9px 22px; background: #334155; color: var(--text-main); border: none; border-radius: 8px; cursor: pointer; font-size: 1rem; font-weight: bold; transition: all 0.2s; }}
  .tab-btn.active {{ background: #3b82f6; box-shadow: 0 4px 10px rgba(59,130,246,0.4); }}
  
  .floor-container {{ background: #111827; border-radius: 14px; padding: 25px; border: 1px solid #334155; display: flex; justify-content: center; overflow-x: auto; }}
  .room-table {{ border-collapse: collapse; background: #1e293b; border: 2px solid #64748b; margin: 0 auto; }}
  .room-table td {{ border: 1px solid #475569; padding: 10px 14px; text-align: center; vertical-align: middle; min-width: 75px; height: 50px; font-size: 1.05rem; }}
  
  .seat-code {{ background: #0f172a; color: #60a5fa; font-weight: 800; width: 65px; border-right: 1px solid #475569; }}
  .seat-name-cell {{ background: #1e293b; color: #f8fafc; font-weight: 700; width: 120px; transition: all 0.3s; }}
  .seat-name-cell.occupied {{ background: rgba(16, 185, 129, 0.2); color: #34d399; }}
  .seat-name-cell.flash {{ animation: flashAnim 1.5s ease-out; }}

  .door-cell {{ background: #334155; color: #f8fafc; font-weight: 800; letter-spacing: 2px; }}
  .feature-cell {{ background: #475569; color: #e2e8f0; font-weight: 800; }}
  .aisle-cell {{ background: #1e293b; color: #94a3b8; font-style: italic; letter-spacing: 8px; font-size: 0.95rem; border: 1px dashed #64748b !important; }}

  @keyframes flashAnim {{
    0% {{ background: #f59e0b; color: #000; transform: scale(1.05); }}
    100% {{ background: rgba(16, 185, 129, 0.2); color: #34d399; transform: scale(1); }}
  }}
</style>
</head>
<body>
<div class="container">
  <header>
    <h1>🎓 2026 新生茶會 ‧ 研究室抽籤開獎</h1>
    <p>現場即時開獎系統 ｜ 🏛️ 100% 擬真房間格局 ‧ 蛇形相鄰不拆組</p>
  </header>

  <div class="room-banner">
    <div class="room-card" id="card-209">
      <div class="room-title"><span>研究室 209</span><span id="stat-209">0 / 20</span></div>
      <div class="room-progress-bg"><div class="room-progress-bar" id="bar-209"></div></div>
      <div class="room-stat" id="rem-209">剩餘 20 席</div>
    </div>
    <div class="room-card" id="card-310">
      <div class="room-title"><span>研究室 310</span><span id="stat-310">0 / 27</span></div>
      <div class="room-progress-bg"><div class="room-progress-bar" id="bar-310"></div></div>
      <div class="room-stat" id="rem-310">剩餘 27 席</div>
    </div>
    <div class="room-card" id="card-313">
      <div class="room-title"><span>研究室 313</span><span id="stat-313">0 / 23</span></div>
      <div class="room-progress-bg"><div class="room-progress-bar" id="bar-313"></div></div>
      <div class="room-stat" id="rem-313">剩餘 23 席</div>
    </div>
    <div class="room-card" id="card-919">
      <div class="room-title"><span>研究室 919</span><span id="stat-919">0 / 9</span></div>
      <div class="room-progress-bg"><div class="room-progress-bar" id="bar-919"></div></div>
      <div class="room-stat" id="rem-919">剩餘 9 席</div>
    </div>
  </div>

  <div class="stage-box">
    <div class="stage-title" id="stage-desc">準備就緒 ‧ 請按下方按鈕開始</div>
    <div class="group-info-header" id="group-target">點擊「開出下一組」依序開獎</div>
    <div class="members-grid-box" id="members-container"></div>
    <div class="result-banner" id="result-banner">???</div>
    <div class="btn-group">
      <button class="btn btn-primary" id="btn-next" onclick="drawNext()">🎲 開出下一組</button>
      <button class="btn btn-accent" id="btn-auto" onclick="toggleAuto()">⚡ 自動連續開獎</button>
    </div>
  </div>

  <div class="floor-section">
    <div class="floor-tabs">
      <button class="tab-btn active" onclick="switchTab('209')">209 平面格局圖 (20席)</button>
      <button class="tab-btn" onclick="switchTab('310')">310 平面格局圖 (27席)</button>
      <button class="tab-btn" onclick="switchTab('313')">313 平面格局圖 (23席)</button>
      <button class="tab-btn" onclick="switchTab('919')">919 平面格局圖 (9席)</button>
    </div>
    <div class="floor-container" id="floor-view"></div>
  </div>
</div>

<script>
const drawHistory = {json.dumps(draw_data, ensure_ascii=False)};
const caps = {{ "209": 20, "310": 27, "313": 23, "919": 9 }};
let counts = {{ "209": 0, "310": 0, "313": 0, "919": 0 }};
let currentIdx = 0;
let currentTab = "209";
let autoInterval = null;
let seatOccupancy = {{ "209": {{}}, "310": {{}}, "313": {{}}, "919": {{}} }};

function renderCurrentRoom() {{
  const container = document.getElementById("floor-view");
  if (currentTab === "209") {{
    let html = `<table class="room-table">
      <tr><td colspan="8" class="door-cell">大 門</td></tr>`;
    for (let r = 1; r <= 5; r++) {{
      html += `<tr>
        <td class="seat-code">1-${{r}}</td><td class="seat-name-cell" id="seat-209-1-${{r}}">${{seatOccupancy['209']['1-'+r] || ''}}</td>
        <td class="seat-code">2-${{r}}</td><td class="seat-name-cell" id="seat-209-2-${{r}}">${{seatOccupancy['209']['2-'+r] || ''}}</td>
        <td class="seat-code">3-${{r}}</td><td class="seat-name-cell" id="seat-209-3-${{r}}">${{seatOccupancy['209']['3-'+r] || ''}}</td>
        <td class="seat-code">4-${{r}}</td><td class="seat-name-cell" id="seat-209-4-${{r}}">${{seatOccupancy['209']['4-'+r] || ''}}</td>
      </tr>`;
    }}
    html += `</table>`;
    container.innerHTML = html;
  }} else if (currentTab === "310") {{
    let html = `<table class="room-table">
      <tr>
        <td colspan="2" class="door-cell">大 門</td>
        <td colspan="10" style="background:transparent; border:none;"></td>
      </tr>`;
    for (let r = 1; r <= 5; r++) {{
      html += `<tr>
        <td class="seat-code">1-${{r}}</td><td class="seat-name-cell" id="seat-310-1-${{r}}">${{seatOccupancy['310']['1-'+r] || ''}}</td>
        <td class="seat-code">2-${{r}}</td><td class="seat-name-cell" id="seat-310-2-${{r}}">${{seatOccupancy['310']['2-'+r] || ''}}</td>
        <td class="seat-code">3-${{r}}</td><td class="seat-name-cell" id="seat-310-3-${{r}}">${{seatOccupancy['310']['3-'+r] || ''}}</td>
        <td class="seat-code">4-${{r}}</td><td class="seat-name-cell" id="seat-310-4-${{r}}">${{seatOccupancy['310']['4-'+r] || ''}}</td>
        <td class="seat-code">5-${{r}}</td><td class="seat-name-cell" id="seat-310-5-${{r}}">${{seatOccupancy['310']['5-'+r] || ''}}</td>`;
      if (r === 1) {{
        html += `<td colspan="2" rowspan="3" class="feature-cell">印<br>表<br>機<br>區</td>`;
      }} else if (r === 4) {{
        html += `<td class="seat-code">6-1</td><td class="seat-name-cell" id="seat-310-6-1">${{seatOccupancy['310']['6-1'] || ''}}</td>`;
      }} else if (r === 5) {{
        html += `<td class="seat-code">6-2</td><td class="seat-name-cell" id="seat-310-6-2">${{seatOccupancy['310']['6-2'] || ''}}</td>`;
      }}
      html += `</tr>`;
    }}
    html += `</table>`;
    container.innerHTML = html;
  }} else if (currentTab === "313") {{
    let html = `<table class="room-table">
      <tr>
        <td colspan="6" style="background:transparent; border:none;"></td>
        <td colspan="2" class="door-cell">大 門</td>
      </tr>`;
    for (let r = 1; r <= 6; r++) {{
      html += `<tr>
        <td class="seat-code">1-${{r}}</td><td class="seat-name-cell" id="seat-313-1-${{r}}">${{seatOccupancy['313']['1-'+r] || ''}}</td>
        <td class="seat-code">2-${{r}}</td><td class="seat-name-cell" id="seat-313-2-${{r}}">${{seatOccupancy['313']['2-'+r] || ''}}</td>
        <td class="seat-code">3-${{r}}</td><td class="seat-name-cell" id="seat-313-3-${{r}}">${{seatOccupancy['313']['3-'+r] || ''}}</td>`;
      if (r === 1) {{
        html += `<td class="seat-code">4-1</td><td class="seat-name-cell" id="seat-313-4-1">${{seatOccupancy['313']['4-1'] || ''}}</td>`;
      }} else if (r === 2) {{
        html += `<td colspan="2" class="feature-cell">牆 柱</td>`;
      }} else {{
        const seatNum = r - 1;
        html += `<td class="seat-code">4-${{seatNum}}</td><td class="seat-name-cell" id="seat-313-4-${{seatNum}}">${{seatOccupancy['313']['4-'+seatNum] || ''}}</td>`;
      }}
      html += `</tr>`;
    }}
    html += `</table>`;
    container.innerHTML = html;
  }} else if (currentTab === "919") {{
    let html = `<table class="room-table">
      <tr>
        <td class="door-cell">大 門</td>
        <td class="feature-cell">公用電腦</td>
        <td class="seat-code">2-1</td>
        <td class="seat-name-cell" id="seat-919-2-1">${{seatOccupancy['919']['2-1'] || ''}}</td>
      </tr>
      <tr>
        <td class="seat-code">1-1</td><td class="seat-name-cell" id="seat-919-1-1">${{seatOccupancy['919']['1-1'] || ''}}</td>
        <td class="seat-code">2-2</td><td class="seat-name-cell" id="seat-919-2-2">${{seatOccupancy['919']['2-2'] || ''}}</td>
      </tr>
      <tr>
        <td class="seat-code">1-2</td><td class="seat-name-cell" id="seat-919-1-2">${{seatOccupancy['919']['1-2'] || ''}}</td>
        <td class="seat-code">2-3</td><td class="seat-name-cell" id="seat-919-2-3">${{seatOccupancy['919']['2-3'] || ''}}</td>
      </tr>
      <tr>
        <td class="seat-code">1-3</td><td class="seat-name-cell" id="seat-919-1-3">${{seatOccupancy['919']['1-3'] || ''}}</td>
        <td class="seat-code">2-4</td><td class="seat-name-cell" id="seat-919-2-4">${{seatOccupancy['919']['2-4'] || ''}}</td>
      </tr>
      <tr>
        <td colspan="4" class="aisle-cell">走 道</td>
      </tr>
      <tr>
        <td class="seat-code">1-4</td><td class="seat-name-cell" id="seat-919-1-4">${{seatOccupancy['919']['1-4'] || ''}}</td>
        <td class="seat-code">2-5</td><td class="seat-name-cell" id="seat-919-2-5">${{seatOccupancy['919']['2-5'] || ''}}</td>
      </tr>
    </table>`;
    container.innerHTML = html;
  }}

  for (let lbl in seatOccupancy[currentTab]) {{
    if (seatOccupancy[currentTab][lbl]) {{
      const el = document.getElementById(`seat-${{currentTab}}-${{lbl}}`);
      if (el) el.classList.add('occupied');
    }}
  }}
}}

function switchTab(lab) {{
  currentTab = lab;
  document.querySelectorAll(".tab-btn").forEach(b => {{
    b.classList.toggle("active", b.innerText.includes(lab));
  }});
  renderCurrentRoom();
}}

function drawNext() {{
  if (currentIdx >= drawHistory.length) return;
  const cur = drawHistory[currentIdx];
  const btnNext = document.getElementById("btn-next");
  btnNext.disabled = true;

  document.getElementById("stage-desc").innerText = `第 [${{cur.order}}/28] 組 開獎中...`;
  document.getElementById("group-target").innerText = `【 第 ${{cur.group_id}} 組 】（共 ${{cur.size}} 人）`;
  document.getElementById("members-container").innerHTML = cur.members.map(m => `<div class="member-large-tag">👤 ${{m}}</div>`).join("");
  
  const banner = document.getElementById("result-banner");
  banner.className = "result-banner";
  const labs = ["209", "310", "313", "919"];
  let roll = 0;
  const timer = setInterval(() => {{
    banner.innerText = labs[roll % labs.length];
    roll++;
    if (roll > 10) {{
      clearInterval(timer);
      banner.innerText = `研究室 ${{cur.lab}}`;
      banner.classList.add("revealed");
      confetti({{ particleCount: 45, spread: 60, origin: {{ y: 0.6 }} }});
      
      counts[cur.lab] += cur.size;
      document.getElementById(`stat-${{cur.lab}}`).innerText = `${{counts[cur.lab]}} / ${{caps[cur.lab]}}`;
      document.getElementById(`bar-${{cur.lab}}`).style.width = `${{(counts[cur.lab]/caps[cur.lab])*100}}%`;
      document.getElementById(`rem-${{cur.lab}}`).innerText = `剩餘 ${{caps[cur.lab] - counts[cur.lab]}} 席`;
      
      cur.members.forEach((m, i) => {{ seatOccupancy[cur.lab][cur.seat_labels[i]] = m; }});
      currentTab = cur.lab;
      document.querySelectorAll(".tab-btn").forEach(b => b.classList.toggle("active", b.innerText.includes(cur.lab)));
      renderCurrentRoom();
      
      cur.seat_labels.forEach(lbl => {{
        const el = document.getElementById(`seat-${{cur.lab}}-${{lbl}}`);
        if (el) {{
          el.classList.add("occupied", "flash");
        }}
      }});

      currentIdx++;
      if (currentIdx >= drawHistory.length) {{
        document.getElementById("stage-desc").innerText = "🎉 全數 28 組開獎完成！";
        confetti({{ particleCount: 150, spread: 100 }});
        btnNext.disabled = true;
      }} else {{
        btnNext.disabled = false;
      }}
    }}
  }}, 60);
}}

function toggleAuto() {{
  const btn = document.getElementById("btn-auto");
  if (autoInterval) {{
    clearInterval(autoInterval);
    autoInterval = null;
    btn.innerText = "⚡ 自動連續開獎";
  }} else {{
    btn.innerText = "⏸ 暫停開獎";
    drawNext();
    autoInterval = setInterval(() => {{
      if (currentIdx >= drawHistory.length) toggleAuto();
      else drawNext();
    }}, 2200);
  }}
}}

renderCurrentRoom();
</script>
</body>
</html>"""

    with open(output_html, "w", encoding="utf-8") as f:
        f.write(html_code)

    print("✅ 實體格局視覺化系統生成完畢！")
    return html_code

# 執行並直接在 Colab 內嵌顯示
html_content = run_real_floorplan_lottery("研究室選位.xlsx")
IPython.display.display(IPython.display.HTML(html_content))